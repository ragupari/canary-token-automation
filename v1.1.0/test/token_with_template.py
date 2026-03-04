import zipfile
import os
import shutil
from openpyxl import load_workbook

# 1. Define paths
original_file = 'generated_tokens/canary_user_02.xlsx'
temp_modified = 'temp_modified.xlsx'
output_file = 'canary_updated.xlsx'

# 2. Create a temporary modified version using your existing code
wb = load_workbook(original_file)
sheet = wb.active
sheet['C1'] = "Updated Value"
sheet.cell(row=5, column=1).value = "New Row Entry"
wb.save(temp_modified)

# 3. Perform the surgical swap
with zipfile.ZipFile(original_file, 'r') as zin:
    with zipfile.ZipFile(output_file, 'w') as zout:
        # Copy everything from original EXCEPT the sheet data
        for item in zin.infolist():
            if item.filename != 'xl/worksheets/sheet1.xml':
                zout.writestr(item, zin.read(item.filename))
        
        # 4. Insert the modified sheet from the temp file
        with zipfile.ZipFile(temp_modified, 'r') as ztemp:
            zout.writestr('xl/worksheets/sheet1.xml', ztemp.read('xl/worksheets/sheet1.xml'))

# 5. Cleanup
os.remove(temp_modified)
print(f"Success! {output_file} created with data updated and token preserved.")